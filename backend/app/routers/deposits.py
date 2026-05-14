from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import List, Optional
from datetime import date
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from ..database import get_db
from .. import models, schemas
from ..auth import get_current_user, require_admin_or_account

router = APIRouter(prefix="/api/deposits", tags=["deposits"])

COLUMNS = ["member_email", "deposit_type", "amount", "deposit_date", "year", "month", "note"]
DEPOSIT_TYPES = ("monthly", "yearly")


@router.post("/", response_model=schemas.DepositOut)
def create_deposit(data: schemas.DepositCreate, db: Session = Depends(get_db), current_user=Depends(require_admin_or_account)):
    deposit = models.Deposit(**data.model_dump(), created_by=current_user.id)
    db.add(deposit)
    db.commit()
    db.refresh(deposit)
    return deposit


@router.get("/", response_model=List[schemas.DepositOut])
def list_deposits(
    user_id: Optional[int] = None,
    year: Optional[int] = None,
    month: Optional[int] = None,
    deposit_type: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user)
):
    query = db.query(models.Deposit)
    if current_user.role == models.RoleEnum.member:
        query = query.filter(models.Deposit.user_id == current_user.id)
    elif user_id:
        query = query.filter(models.Deposit.user_id == user_id)
    if year:
        query = query.filter(models.Deposit.year == year)
    if month:
        query = query.filter(models.Deposit.month == month)
    if deposit_type:
        query = query.filter(models.Deposit.deposit_type == deposit_type)
    return query.order_by(models.Deposit.deposit_date.desc()).all()


@router.delete("/{deposit_id}")
def delete_deposit(deposit_id: int, db: Session = Depends(get_db), _=Depends(require_admin_or_account)):
    deposit = db.query(models.Deposit).filter(models.Deposit.id == deposit_id).first()
    if not deposit:
        raise HTTPException(status_code=404, detail="Deposit not found")
    db.delete(deposit)
    db.commit()
    return {"message": "Deleted"}


@router.get("/bulk-template")
def download_template(db: Session = Depends(get_db), _=Depends(require_admin_or_account)):
    """Generate and return a demo Excel template with member emails pre-filled."""
    members = db.query(models.User).filter(
        models.User.role == models.RoleEnum.member,
        models.User.is_active == True
    ).all()

    wb = openpyxl.Workbook()

    # ── Instructions sheet ──
    ws_info = wb.active
    ws_info.title = "Instructions"
    ws_info.column_dimensions["A"].width = 22
    ws_info.column_dimensions["B"].width = 55

    header_fill = PatternFill("solid", fgColor="1A3C5E")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    title_font = Font(color="1A3C5E", bold=True, size=14)
    note_fill = PatternFill("solid", fgColor="FFF9E6")
    ok_fill = PatternFill("solid", fgColor="E8F8F0")

    ws_info["A1"] = "UND Bulk Deposit Upload — Instructions"
    ws_info["A1"].font = title_font
    ws_info.merge_cells("A1:B1")
    ws_info.row_dimensions[1].height = 28

    instructions = [
        ("Column", "Description / Rules"),
        ("member_email", "Member's registered email address (must exist in the system)"),
        ("deposit_type", "Either  monthly  or  yearly  (lowercase)"),
        ("amount", "Numeric value only, e.g. 1000 or 5000.50"),
        ("deposit_date", "Date in YYYY-MM-DD format, e.g. 2024-06-15"),
        ("year", "4-digit year the deposit belongs to, e.g. 2024"),
        ("month", "Month number 1–12 (required for monthly, leave blank for yearly)"),
        ("note", "Optional free-text note"),
    ]

    for r, (col, desc) in enumerate(instructions, start=2):
        ws_info.cell(r, 1, col)
        ws_info.cell(r, 2, desc)
        if r == 2:
            ws_info.cell(r, 1).font = header_font
            ws_info.cell(r, 1).fill = header_fill
            ws_info.cell(r, 2).font = header_font
            ws_info.cell(r, 2).fill = header_fill
        else:
            ws_info.cell(r, 1).fill = note_fill
            ws_info.cell(r, 2).fill = note_fill
        ws_info.row_dimensions[r].height = 20

    ws_info["A11"] = "⚠ Important Notes"
    ws_info["A11"].font = Font(bold=True, color="C0392B")
    ws_info.merge_cells("A11:B11")

    notes = [
        "• Do NOT change column headers in the Deposits sheet.",
        "• Rows with errors will be skipped — check the upload result summary.",
        "• Duplicate deposits (same member + type + year + month) will still be inserted.",
        "• The Deposits sheet already contains sample rows — replace them with real data.",
    ]
    for i, n in enumerate(notes, start=12):
        ws_info.cell(i, 1, n)
        ws_info.merge_cells(f"A{i}:B{i}")
        ws_info.cell(i, 1).fill = ok_fill
        ws_info.row_dimensions[i].height = 18

    # ── Deposits sheet ──
    ws = wb.create_sheet("Deposits")

    col_widths = [28, 14, 12, 16, 8, 8, 24]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Header row
    for c, col in enumerate(COLUMNS, 1):
        cell = ws.cell(1, c, col)
        cell.font = Font(color="FFFFFF", bold=True, size=11)
        cell.fill = PatternFill("solid", fgColor="1A3C5E")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    ws.row_dimensions[1].height = 24
    ws.freeze_panes = "A2"

    today = date.today()

    # Sample rows using real member emails
    sample_rows = []
    for m in members[:5]:  # up to 5 members as demo rows
        sample_rows.append([
            m.email, "monthly", 1000,
            f"{today.year}-{today.month:02d}-01",
            today.year, today.month, "Monthly deposit"
        ])
    # Add a yearly sample
    if members:
        sample_rows.append([
            members[0].email, "yearly", 10000,
            f"{today.year}-01-01",
            today.year, "", "Annual deposit"
        ])
    # If no members yet, add generic placeholder rows
    if not sample_rows:
        sample_rows = [
            ["member@example.com", "monthly", 1000, f"{today.year}-{today.month:02d}-01", today.year, today.month, "Monthly deposit"],
            ["member@example.com", "yearly", 10000, f"{today.year}-01-01", today.year, "", "Annual deposit"],
        ]

    row_fills = [PatternFill("solid", fgColor="F0F7FF"), PatternFill("solid", fgColor="FFFFFF")]
    for r, row in enumerate(sample_rows, 2):
        fill = row_fills[r % 2]
        for c, val in enumerate(row, 1):
            cell = ws.cell(r, c, val)
            cell.fill = fill
            cell.border = border
            cell.alignment = Alignment(vertical="center")
        ws.row_dimensions[r].height = 20

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=UND_Deposit_Upload_Template.xlsx"}
    )


@router.post("/bulk-upload")
def bulk_upload(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user=Depends(require_admin_or_account)
):
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Only .xlsx or .xls files are accepted")

    try:
        contents = file.file.read()
        wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="Could not read the Excel file. Make sure it is a valid .xlsx file.")

    # Support both sheet names
    sheet_name = "Deposits" if "Deposits" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet_name]

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="The sheet is empty.")

    # Validate header
    header = [str(h).strip().lower() if h else "" for h in rows[0]]
    missing = [c for c in COLUMNS if c not in header]
    if missing:
        raise HTTPException(status_code=400, detail=f"Missing columns: {', '.join(missing)}")

    col_idx = {c: header.index(c) for c in COLUMNS}

    # Cache member emails → id
    members = {u.email.lower(): u.id for u in db.query(models.User).filter(models.User.role == models.RoleEnum.member).all()}

    inserted, skipped = 0, []

    for row_num, row in enumerate(rows[1:], start=2):
        def get(col):
            v = row[col_idx[col]]
            return str(v).strip() if v is not None else ""

        email = get("member_email").lower()
        dep_type = get("deposit_type").lower()
        amount_raw = get("amount")
        dep_date_raw = get("deposit_date")
        year_raw = get("year")
        month_raw = get("month")
        note = get("note")

        # Skip blank rows
        if not any([email, dep_type, amount_raw]):
            continue

        errors = []

        if email not in members:
            errors.append(f"email '{email}' not found")
        if dep_type not in DEPOSIT_TYPES:
            errors.append(f"deposit_type must be 'monthly' or 'yearly'")

        try:
            amount = float(amount_raw)
            if amount <= 0:
                raise ValueError
        except (ValueError, TypeError):
            errors.append("amount must be a positive number")
            amount = None

        try:
            dep_date = date.fromisoformat(dep_date_raw)
        except (ValueError, TypeError):
            errors.append("deposit_date must be YYYY-MM-DD")
            dep_date = None

        try:
            year = int(float(year_raw))
        except (ValueError, TypeError):
            errors.append("year must be a number")
            year = None

        month = None
        if dep_type == "monthly":
            try:
                month = int(float(month_raw))
                if not 1 <= month <= 12:
                    raise ValueError
            except (ValueError, TypeError):
                errors.append("month must be 1–12 for monthly deposits")
                month = None

        if errors:
            skipped.append({"row": row_num, "email": email, "errors": errors})
            continue

        deposit = models.Deposit(
            user_id=members[email],
            deposit_type=dep_type,
            amount=amount,
            deposit_date=dep_date,
            year=year,
            month=month,
            note=note or None,
            created_by=current_user.id
        )
        db.add(deposit)
        inserted += 1

    db.commit()
    return {
        "inserted": inserted,
        "skipped": len(skipped),
        "skipped_details": skipped
    }
